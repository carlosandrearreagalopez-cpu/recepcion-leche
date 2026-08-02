import streamlit as st
import pandas as pd
import os
from datetime import datetime
from streamlit_drawable_canvas import st_canvas
import io
from PIL import Image

# Importaciones para Excel
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Importaciones para PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import Table as RLTable, TableStyle, Image as RLImage, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# ==========================================
# CONFIGURACIÓN Y ESTILOS (Colores LIF Brands)
# ==========================================
st.set_page_config(page_title="Control de Recepción - LIF Brands", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
.stApp { background-color: #F8FAF9 !important; }
html, body, p, label, div { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #1f2937; }
h1, h2, h3, h4, h5, h6 { color: #1e3a8a !important; font-weight: 700 !important; }

.stTextInput label, .stSelectbox label, .stDateInput label, .stNumberInput label, .stFileUploader label, .stTextArea label, .stRadio label {
    color: #1e3a8a !important;
    font-weight: bold !important;
}

[data-testid="stExpander"] details summary {
    background-color: #f1f5f9 !important;
    color: #1e3a8a !important;
    font-weight: bold;
    border-radius: 6px;
}
[data-testid="stExpander"] details summary p { color: #1e3a8a !important; font-weight: bold; }

[data-testid="baseButton-primary"] {
    background-color: #1e3a8a !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: bold !important;
}
[data-testid="baseButton-primary"]:hover { background-color: #1e40af !important; }

.record-card {
    background-color: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    padding: 16px;
    margin-bottom: 5px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    border-left: 5px solid #3b82f6;
}
.record-header { font-size: 1.1em; font-weight: bold; color: #1e3a8a; margin-bottom: 8px; }
.record-sub { color: #475569; font-size: 0.9em; }
.status-pendiente { color: #d97706; font-weight: bold; }
.status-aprobado { color: #16a34a; font-weight: bold; }
.status-rechazado { color: #dc2626; font-weight: bold; }

canvas.lower-canvas, canvas.upper-canvas { border: 1px solid #cbd5e1 !important; border-radius: 4px !important; }
</style>
""", unsafe_allow_html=True)

# ==========================================
# DIRECTORIOS Y FUNCIONES BASE
# ==========================================
FIRMAS_DIR = "firmas_recepcion"
FIRMAS_REGISTRADAS_DIR = "firmas_registradas_jefes"
EVIDENCIAS_DIR = "evidencias_recepcion"

for directorio in [FIRMAS_DIR, FIRMAS_REGISTRADAS_DIR, EVIDENCIAS_DIR]:
    if not os.path.exists(directorio):
        os.makedirs(directorio)

EXCEL_FILE = "registros_recepcion_leche.xlsx"

def mostrar_logo(ancho=160):
    if os.path.exists("logo.png"):
        st.image("logo.png", width=ancho)

def cargar_datos():
    if os.path.exists(EXCEL_FILE):
        df = pd.read_excel(EXCEL_FILE, dtype={"ID_Registro": str})
        columnas_requeridas = {
            "ID_Registro": "", "Estado": "Pendiente", "Responsable": "", "Fecha": "", "Hora": "",
            "Proveedor": "", "Cantidad_Litros": 0.0, "Limpieza_Exterior": "", "Salidas_Selladas": "", 
            "Desinfeccion_Utensilios": "", "Temperatura_C": 0.0, "Color": "", "Olor": "", "Sabor": "", 
            "Apariencia": "", "pH": 0.0, "Acido_Lactico": 0.0, "Grasa": 0.0, "Solido_No_Graso": 0.0,
            "Solido_Total": 0.0, "Densidad": 0.0, "Punto_Congelacion": 0.0, "Proteina": 0.0, 
            "Lactosa": 0.0, "Conductividad": 0.0, "Agua_Anadida": 0.0, "Antibioticos": "", 
            "Peroxido": "", "Carga_Adecuada": "", "Afecto_Ambiente": "", "Resolucion": "",
            "Evidencia": "", "Firma_Jefe": "Sin firma", "Observaciones_Jefe": ""
        }
        for col, val_default in columnas_requeridas.items():
            if col not in df.columns:
                df[col] = val_default
        
        for col in ["Estado", "Firma_Jefe", "Observaciones_Jefe", "Evidencia"]:
            df[col] = df[col].astype(str)
        return df
    return pd.DataFrame()

def guardar_datos(df):
    df.to_excel(EXCEL_FILE, index=False)

def eliminar_registro(id_registro):
    df = cargar_datos()
    df = df[df["ID_Registro"] != str(id_registro)]
    guardar_datos(df)
    st.success(f"Registro #{id_registro} eliminado.")

def generar_id_registro():
    fecha_base = datetime.today().strftime("%Y%m%d")
    df = cargar_datos()
    if df.empty or "ID_Registro" not in df.columns: return fecha_base
    ids_existentes = df["ID_Registro"].astype(str).tolist()
    if fecha_base not in ids_existentes: return fecha_base
    contador = 1
    while f"{fecha_base}({contador})" in ids_existentes:
        contador += 1
    return f"{fecha_base}({contador})"

def generar_excel_bytes(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Registros')
        worksheet = writer.sheets['Registros']
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        if max_row > 1:
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            tab = Table(displayName="TablaRegistros", ref=ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            worksheet.add_table(tab)
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(cell.value)
                    except: pass
                worksheet.column_dimensions[column].width = min(max_length + 2, 35)
    return output.getvalue()

# ==========================================
# GENERADOR DE PDF (Formato Oficial de Leche)
# ==========================================
def generar_pdf_nuevo(registro):
    buffer = io.BytesIO()
    c = pdf_canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    margin_x, margin_y = 30, 40
    usable_width = width - (2 * margin_x)
    col_widths = [usable_width * 0.35, usable_width * 0.15, usable_width * 0.25, usable_width * 0.25]
    styles = getSampleStyleSheet()
    
    style_normal = ParagraphStyle('CellNormal', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, textColor=colors.black)
    style_bold = ParagraphStyle('CellBold', parent=style_normal, fontName='Helvetica-Bold')
    style_center_bold = ParagraphStyle('CellCenterBold', parent=style_bold, alignment=1)
    style_meta = ParagraphStyle('CellMeta', parent=style_normal, fontSize=7, leading=9, alignment=1)

    logo_img = RLImage("logo.png", width=70, height=25) if os.path.exists("logo.png") else ""
        
    data = [
        [logo_img, Paragraph("RECEPCIÓN DE LECHE CRUDA", style_center_bold), "", Paragraph(f"Código: R PBD/01-1<br/>Versión: 8<br/>ID: #{registro.get('ID_Registro', '')}", style_meta)],
        [Paragraph("Fecha de recepción:", style_bold), Paragraph(str(registro.get('Fecha', '')), style_normal), Paragraph("Proveedor:", style_bold), Paragraph(str(registro.get('Proveedor', '')), style_normal)],
        [Paragraph("Cantidad reportada (L):", style_bold), Paragraph(str(registro.get('Cantidad_Litros', '')), style_normal), "", ""],
        
        [Paragraph("Vehículo de transporte", style_center_bold), Paragraph("Resultado", style_center_bold), Paragraph("Observaciones de Calidad", style_center_bold), ""],
        [Paragraph("Limpieza exterior", style_normal), Paragraph(str(registro.get('Limpieza_Exterior', '')), style_normal), Paragraph(str(registro.get('Observaciones_Jefe', '')), style_normal), ""],
        [Paragraph("Salidas de leche selladas", style_normal), Paragraph(str(registro.get('Salidas_Selladas', '')), style_normal), "", ""],
        [Paragraph("Desinfección utensilios", style_normal), Paragraph(str(registro.get('Desinfeccion_Utensilios', '')), style_normal), "", ""],
        
        [Paragraph("Análisis Fisicoquímico", style_center_bold), Paragraph("Resultado", style_center_bold), Paragraph("Análisis Fisicoquímico", style_center_bold), Paragraph("Resultado", style_center_bold)],
        [Paragraph("Temperatura (≤ 7°C)", style_normal), Paragraph(str(registro.get('Temperatura_C', '')), style_normal), Paragraph("% Ácido láctico (0.13-0.17)", style_normal), Paragraph(str(registro.get('Acido_Lactico', '')), style_normal)],
        [Paragraph("Color (blanco, crema claro)", style_normal), Paragraph(str(registro.get('Color', '')), style_normal), Paragraph("Antibióticos (negativo)", style_normal), Paragraph(str(registro.get('Antibioticos', '')), style_normal)],
        [Paragraph("Olor (característico)", style_normal), Paragraph(str(registro.get('Olor', '')), style_normal), Paragraph("Peróxido (negativo)", style_normal), Paragraph(str(registro.get('Peroxido', '')), style_normal)],
        [Paragraph("Sabor (característico)", style_normal), Paragraph(str(registro.get('Sabor', '')), style_normal), Paragraph("% Grasa (mín. 3%)", style_normal), Paragraph(str(registro.get('Grasa', '')), style_normal)],
        [Paragraph("Apariencia", style_normal), Paragraph(str(registro.get('Apariencia', '')), style_normal), Paragraph("% Sólido No Graso (mín. 8.3%)", style_normal), Paragraph(str(registro.get('Solido_No_Graso', '')), style_normal)],
        [Paragraph("pH (6.5 - 6.8)", style_normal), Paragraph(str(registro.get('pH', '')), style_normal), Paragraph("% Sólido Total (mín. 11.3%)", style_normal), Paragraph(str(registro.get('Solido_Total', '')), style_normal)],
        [Paragraph("Densidad (1.030 - 1.033 g/ml)", style_normal), Paragraph(str(registro.get('Densidad', '')), style_normal), Paragraph("Pto. congelación", style_normal), Paragraph(str(registro.get('Punto_Congelacion', '')), style_normal)],
        [Paragraph("% Proteína (mín. 3%)", style_normal), Paragraph(str(registro.get('Proteina', '')), style_normal), Paragraph("Lactosa (mín. 4.2%)", style_normal), Paragraph(str(registro.get('Lactosa', '')), style_normal)],
        [Paragraph("Conductividad (máx. 6mS/cm)", style_normal), Paragraph(str(registro.get('Conductividad', '')), style_normal), Paragraph("Agua Añadida (0%)", style_normal), Paragraph(str(registro.get('Agua_Anadida', '')), style_normal)],
        
        [Paragraph("Resolución y Recepción de Leche", style_center_bold), "", "", ""],
        [Paragraph("Estado Final:", style_bold), Paragraph(str(registro.get('Estado', '')), style_bold), Paragraph("Resolución de Planta:", style_bold), Paragraph(str(registro.get('Resolucion', '')), style_normal)],
        [Paragraph("¿El proceso de carga/descarga fue adecuado?", style_normal), Paragraph(str(registro.get('Carga_Adecuada', '')), style_normal), Paragraph("¿Afectó de forma potencial al medio ambiente?", style_normal), Paragraph(str(registro.get('Afecto_Ambiente', '')), style_normal)],
        [Paragraph("Realizado por:", style_bold), Paragraph(str(registro.get('Responsable', '')), style_normal), Paragraph("Verificado por:", style_bold), ""],
    ]
    
    style = TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black), 
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        
        # Header Spans
        ('SPAN', (1,0), (2,0)), 
        ('SPAN', (2,2), (3,2)), 
        
        # Transporte Spans
        ('SPAN', (0,3), (3,3)), ('BACKGROUND', (0,3), (3,3), colors.lightgrey),
        ('SPAN', (2,4), (3,6)), ('VALIGN', (2,4), (3,6), 'TOP'),
        
        # FQ Spans
        ('SPAN', (0,7), (3,7)), ('BACKGROUND', (0,7), (3,7), colors.lightgrey),
        
        # Resolucion Spans
        ('SPAN', (0,17), (3,17)), ('BACKGROUND', (0,17), (3,17), colors.lightgrey),
        
        # Firmas span area
        ('SPAN', (3,20), (3,20)), 
    ])
    
    # Alturas de fila adaptadas
    row_heights = [35] + [20]*2 + [18] + [25]*3 + [18] + [22]*9 + [18] + [25]*2 + [40]
    t = RLTable(data, colWidths=col_widths, rowHeights=row_heights)
    t.setStyle(style)
    
    w, h = t.wrapOn(c, usable_width, height)
    y_pos_table = height - margin_y - h
    t.drawOn(c, margin_x, y_pos_table)
    
    # Colocar la firma del Jefe de Calidad en la última celda ("Verificado por")
    nombre_firma = registro.get("Firma_Jefe", "Sin firma")
    ruta_firma = os.path.join(FIRMAS_DIR, str(nombre_firma))
    
    if nombre_firma != "Sin firma" and os.path.exists(ruta_firma):
        # Coordenadas relativas a la tabla (Columna 3, Fila 20)
        x_firma = margin_x + col_widths[0] + col_widths[1] + col_widths[2] + 10
        y_firma = y_pos_table + 5 # Posición bottom de la tabla
        c.drawImage(ruta_firma, x_firma, y_firma, width=100, height=35, preserveAspectRatio=True, mask='auto')
    
    c.save()
    buffer.seek(0)
    return buffer.getvalue()

# ==========================================
# CONTROL DE ESTADOS DE SESIÓN
# ==========================================
for state in ["nav_state", "form_logueado", "admin_logueado", "enviado_exitoso"]:
    if state not in st.session_state:
        st.session_state[state] = "home" if state == "nav_state" else False

# ==========================================
# 1. PANTALLA DE INICIO
# ==========================================
if st.session_state["nav_state"] == "home":
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        mostrar_logo(200)
        st.markdown("<h1 style='text-align: center;'>Registro de Recepción de Leche Cruda</h1>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; font-weight: bold;'>LIF Brands - Aseguramiento de Calidad</p>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        
        if st.button("📝 Colaborador: Reportar nuevo ingreso", use_container_width=True, type="primary"):
            st.session_state["nav_state"] = "form_login"
            st.rerun()
            
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔒 Jefe de Calidad: Panel de Administración", use_container_width=True):
            st.session_state["nav_state"] = "admin_login"
            st.rerun()

# ==========================================
# 2. LOGIN (AMBOS PERFILES)
# ==========================================
elif st.session_state["nav_state"] in ["form_login", "admin_login"]:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("⬅️ Volver al inicio"):
            st.session_state["nav_state"] = "home"
            st.rerun()
        
        es_admin = st.session_state["nav_state"] == "admin_login"
        st.title("Panel de Administrador" if es_admin else "Acceso a Registro")
        st.markdown("Ingrese la contraseña autorizada:")
        
        pwd = st.text_input("Contraseña", type="password")
        pwd_correcta = "glad726lif" if es_admin else "20lf26"
        
        if st.button("Verificar Acceso", use_container_width=True, type="primary"):
            if pwd == pwd_correcta:
                st.session_state["admin_logueado" if es_admin else "form_logueado"] = True
                st.session_state["nav_state"] = "admin_dashboard" if es_admin else "form"
                st.rerun()
            else:
                st.error("Contraseña incorrecta.")

# ==========================================
# 3. FORMULARIO DEL COLABORADOR
# ==========================================
elif st.session_state["nav_state"] == "form":
    if not st.session_state["form_logueado"]:
        st.session_state["nav_state"] = "home"
        st.rerun()
        
    c_h1, c_h2 = st.columns([4, 1])
    with c_h2:
        if st.button("⬅️ Cerrar Sesión", use_container_width=True):
            st.session_state["form_logueado"] = False
            st.session_state["nav_state"] = "home"
            st.rerun()
    
    mostrar_logo(140)
    st.title("Recepción de Leche Cruda")
    
    if st.session_state["enviado_exitoso"]:
        st.success("¡Registro enviado con éxito! Quedará pendiente de validación.")
        if st.button("➕ Ingresar un nuevo registro", type="primary"):
            st.session_state["enviado_exitoso"] = False
            st.rerun()
    else:
        with st.form("form_leche"):
            st.header("1. Datos Generales")
            c1, c2 = st.columns(2)
            with c1:
                responsable = st.selectbox("Nombre del responsable", ["Sandra Garcia", "Daniel Castro", "Luis Perez", "Carlos López", "Marlon Escobar"])
                proveedor_opcion = st.selectbox("Proveedor", ["Pasajinak", "Otro"])
                proveedor_final = st.text_input("Especifique el nombre:") if proveedor_opcion == "Otro" else proveedor_opcion
            with c2:
                fecha = st.date_input("Fecha de recepción", value=datetime.today())
                hora = st.text_input("Hora (HH:MM)", value=datetime.now().strftime("%H:%M"))
                cantidad_leche = st.number_input("Cantidad reportada (Litros)", min_value=0.0, value=0.0)
            
            st.header("2. Vehículo de Transporte")
            cv1, cv2, cv3 = st.columns(3)
            with cv1: limpieza = st.radio("Limpieza Exterior", ["Bueno", "Malo"], horizontal=True)
            with cv2: salidas = st.radio("Salidas Selladas", ["Bueno", "Malo"], horizontal=True)
            with cv3: desinfeccion = st.radio("Desinfección utensilios", ["Bueno", "Malo"], horizontal=True)

            st.header("3. Análisis Físicoquímico")
            temp = st.number_input("Temperatura (°C)", format="%.2f")
            
            cf1, cf2, cf3 = st.columns(3)
            with cf1:
                color = st.radio("Color", ["Característico", "No característico"], horizontal=True)
                olor = st.radio("Olor", ["Característico", "No característico"], horizontal=True)
                sabor = st.radio("Sabor", ["Característico", "No característico"], horizontal=True)
                apariencia = st.radio("Apariencia", ["Sin coágulos", "Con coágulos"], horizontal=True)
                ph = st.number_input("pH", format="%.2f")
                grasa = st.number_input("% Grasa", format="%.2f")
            with cf2:
                densidad = st.number_input("Densidad", format="%.4f")
                lactosa = st.number_input("Lactosa", format="%.2f")
                antibioticos = st.radio("Antibióticos", ["Negativo", "Positivo"], horizontal=True)
                acido_lactico = st.number_input("% Ácido láctico", format="%.4f")
                sng = st.number_input("% SNG", format="%.2f")
                congelacion = st.number_input("Congelación", format="%.4f")
            with cf3:
                conductividad = st.number_input("Conductividad", format="%.2f")
                peroxido = st.radio("Peróxido", ["Negativo", "Positivo"], horizontal=True)
                st_val = st.number_input("% Sólido Total", format="%.2f")
                proteina = st.number_input("% Proteína", format="%.2f")
                agua = st.number_input("% Agua Añadida", format="%.2f")

            st.header("4. Resolución y Evidencia")
            cr1, cr2, cr3 = st.columns(3)
            with cr1: adecuado = st.radio("Proceso adecuado", ["Si", "No"], horizontal=True)
            with cr2: afecto = st.radio("Afectó ambiente", ["Si", "No"], horizontal=True)
            with cr3: resolucion = st.radio("Resolución Planta", ["Aceptado", "Rechazado"], horizontal=True)

            evidencia_foto = st.file_uploader("Evidencia fotográfica (Antibióticos u otros)", type=["png", "jpg", "jpeg"])

            submitted = st.form_submit_button("Guardar y Enviar a Revisión", type="primary")
            
            if submitted:
                id_nuevo = generar_id_registro()
                nombre_evidencia = ""
                if evidencia_foto is not None:
                    img = Image.open(evidencia_foto)
                    if img.mode != 'RGB': img = img.convert('RGB')
                    nombre_evidencia = f"evidencia_{id_nuevo}.jpg"
                    img.save(os.path.join(EVIDENCIAS_DIR, nombre_evidencia))

                nuevo_registro = {
                    "ID_Registro": str(id_nuevo), "Estado": "Pendiente",
                    "Responsable": str(responsable), "Fecha": str(fecha), "Hora": str(hora),
                    "Proveedor": str(proveedor_final), "Cantidad_Litros": float(cantidad_leche),
                    "Limpieza_Exterior": str(limpieza), "Salidas_Selladas": str(salidas), "Desinfeccion_Utensilios": str(desinfeccion),
                    "Temperatura_C": float(temp), "Color": str(color), "Olor": str(olor), "Sabor": str(sabor), "Apariencia": str(apariencia),
                    "pH": float(ph), "Acido_Lactico": float(acido_lactico), "Grasa": float(grasa), "Solido_No_Graso": float(sng),
                    "Solido_Total": float(st_val), "Densidad": float(densidad), "Punto_Congelacion": float(congelacion),
                    "Proteina": float(proteina), "Lactosa": float(lactosa), "Conductividad": float(conductividad), 
                    "Agua_Anadida": float(agua), "Antibioticos": str(antibioticos), "Peroxido": str(peroxido),
                    "Carga_Adecuada": str(adecuado), "Afecto_Ambiente": str(afecto), "Resolucion": str(resolucion),
                    "Evidencia": str(nombre_evidencia), "Firma_Jefe": "Sin firma", "Observaciones_Jefe": ""
                }
                df = cargar_datos()
                df = pd.concat([df, pd.DataFrame([nuevo_registro])], ignore_index=True)
                guardar_datos(df)
                st.session_state["enviado_exitoso"] = True
                st.rerun()

# ==========================================
# 4. DASHBOARD DEL ADMINISTRADOR
# ==========================================
elif st.session_state["nav_state"] == "admin_dashboard":
    if not st.session_state.get("admin_logueado", False):
        st.session_state["nav_state"] = "home"
        st.rerun()
        
    c_head1, c_head2 = st.columns([5, 1])
    with c_head1: st.title("Panel de Administrador - Jefe de Calidad")
    with c_head2:
        if st.button("Cerrar sesión"):
            st.session_state["admin_logueado"] = False
            st.session_state["nav_state"] = "home"
            st.rerun()
            
    df = cargar_datos()
    
    total_pen = len(df[df["Estado"] == "Pendiente"]) if not df.empty else 0
    total_apr = len(df[df["Estado"] == "Aprobado"]) if not df.empty else 0
    total_rec = len(df[df["Estado"] == "Rechazado"]) if not df.empty else 0
    total_reg = len(df) if not df.empty else 0
    
    tab_pendientes, tab_aprobados, tab_rechazados, tab_todos = st.tabs([
        f"⏳ Pendientes ({total_pen})", f"✅ Aprobados ({total_apr})", f"❌ Rechazados ({total_rec})", f"📊 Historial ({total_reg})"
    ])
    
    def render_tarjeta(row, index_key, allow_review=False):
        estado_icono = "⏳" if row['Estado'] == "Pendiente" else "✅" if row['Estado'] == "Aprobado" else "❌"
        css_class = f"status-{str(row['Estado']).lower()}"
        
        st.markdown(f"""
        <div class="record-card">
            <div class="record-header">{estado_icono} #{row['ID_Registro']} — {row['Proveedor']}</div>
            <div class="record-sub">Fecha: {row['Fecha']} | Realizado por: {row['Responsable']} | Litros: {row['Cantidad_Litros']} | Estado: <span class="{css_class}">{row['Estado']}</span></div>
        </div>
        """, unsafe_allow_html=True)
        
        c_btn1, c_btn2, c_btn3 = st.columns([2, 2, 8])
        
        with c_btn1:
            if st.button("🗑️ Eliminar", key=f"del_{index_key}_{row['ID_Registro']}"):
                eliminar_registro(row['ID_Registro'])
                st.rerun()
                
        with c_btn2:
            if row['Estado'] == 'Aprobado':
                pdf_bytes = generar_pdf_nuevo(row.to_dict())
                st.download_button("📥 Formato PDF", data=pdf_bytes, file_name=f"Recepcion_Leche_{row['ID_Registro']}.pdf", mime="application/pdf", key=f"pdf_{index_key}_{row['ID_Registro']}")

        espaciador = " " * (1 if index_key == "pen" else 2 if index_key == "apr" else 3 if index_key == "rec" else 4)
        with st.expander(f"Ver detalles de Análisis #{row['ID_Registro']}{espaciador}"):
            st.write(f"**Temperatura:** {row['Temperatura_C']} °C | **pH:** {row['pH']} | **Densidad:** {row['Densidad']} | **Acidez:** {row['Acido_Lactico']}")
            st.write(f"**Resolución:** {row['Resolucion']} | **Limpieza:** {row['Limpieza_Exterior']} | **Antibióticos:** {row['Antibioticos']}")
            st.write("---")
            
            if "Evidencia" in row and pd.notna(row["Evidencia"]) and row["Evidencia"] != "":
                ruta_evidencia = os.path.join(EVIDENCIAS_DIR, str(row["Evidencia"]))
                if os.path.exists(ruta_evidencia):
                    st.write("**Evidencia Fotográfica adjunta:**")
                    st.image(ruta_evidencia, width=350)
            
            if pd.notna(row.get("Observaciones_Jefe", "")) and str(row.get("Observaciones_Jefe", "")) != "":
                st.warning(f"**Observaciones de Calidad:** {row['Observaciones_Jefe']}")
            
            if allow_review and row['Estado'] == "Pendiente":
                st.markdown("---")
                st.markdown("### Validación y Firma (Jefe de Calidad)")
                
                nombre_jefe = "Jefe de Calidad"
                safe_name = nombre_jefe.replace(" ", "_").lower()
                firma_path_guardada = os.path.join(FIRMAS_REGISTRADAS_DIR, f"firma_reg_{safe_name}.png")
                tiene_firma_previa = os.path.exists(firma_path_guardada)

                modo_firma = "Usar firma guardada"
                if tiene_firma_previa:
                    st.info(f"Firma registrada previamente.")
                    st.image(firma_path_guardada, width=200, caption=f"Firma actual")
                    modo_firma = st.radio("Seleccione opción de firma", ["Usar firma guardada", "Dibujar nueva firma"], horizontal=True, key=f"radio_firma_{row['ID_Registro']}")
                else:
                    st.warning("No hay firma guardada. Dibújela abajo (se guardará para futuros registros).")
                    modo_firma = "Dibujar nueva firma"

                canvas_result = None
                if modo_firma == "Dibujar nueva firma":
                    canvas_result = st_canvas(fill_color="rgba(101, 163, 13, 0.3)", stroke_width=2, stroke_color="#1e3a8a", background_color="#FFFFFF", height=120, width=400, drawing_mode="freedraw", key=f"canvas_firma_{row['ID_Registro']}")

                obs_jefe = st.text_area("Añadir observaciones de calidad (Opcional):", key=f"obs_jefe_{row['ID_Registro']}")
                st.markdown("<br>", unsafe_allow_html=True)
                
                c_rev1, c_rev2 = st.columns(2)
                with c_rev1:
                    if st.button("✅ Aprobar Registro", key=f"btn_aprobar_{row['ID_Registro']}", type="primary"):
                        nombre_firma_archivo = f"firma_{row['ID_Registro']}.png"
                        ruta_destino = os.path.join(FIRMAS_DIR, nombre_firma_archivo)
                        
                        if modo_firma == "Usar firma guardada" and tiene_firma_previa:
                            img_previa = Image.open(firma_path_guardada)
                            if img_previa.mode != 'RGBA': img_previa = img_previa.convert('RGBA')
                            img_previa.save(ruta_destino)
                        else:
                            if canvas_result is not None and canvas_result.image_data is not None:
                                img = Image.fromarray(canvas_result.image_data.astype('uint8'), mode="RGBA")
                                img.save(ruta_destino)
                                img.save(firma_path_guardada)
                            else:
                                st.error("Debe proporcionar una firma válida.")
                                st.stop()

                        df_act = cargar_datos()
                        df_act.loc[df_act['ID_Registro'] == str(row['ID_Registro']), 'Estado'] = "Aprobado"
                        df_act.loc[df_act['ID_Registro'] == str(row['ID_Registro']), 'Firma_Jefe'] = str(nombre_firma_archivo)
                        df_act.loc[df_act['ID_Registro'] == str(row['ID_Registro']), 'Observaciones_Jefe'] = str(obs_jefe)
                        guardar_datos(df_act)
                        st.success("Registro Aprobado exitosamente.")
                        st.rerun()
                
                with c_rev2:
                    if st.button("❌ Rechazar Registro", key=f"btn_rechazar_{row['ID_Registro']}"):
                        df_act = cargar_datos()
                        df_act.loc[df_act['ID_Registro'] == str(row['ID_Registro']), 'Estado'] = "Rechazado"
                        df_act.loc[df_act['ID_Registro'] == str(row['ID_Registro']), 'Observaciones_Jefe'] = str(obs_jefe)
                        guardar_datos(df_act)
                        st.warning("Registro Rechazado.")
                        st.rerun()

    # --- PESTAÑA: PENDIENTES ---
    with tab_pendientes:
        if not df.empty:
            df_pen = df[df["Estado"] == "Pendiente"]
            for idx, row in df_pen.iterrows(): render_tarjeta(row, "pen", allow_review=True)
        else: st.info("No hay registros pendientes.")

    # --- PESTAÑA: APROBADOS ---
    with tab_aprobados:
        if not df.empty:
            df_apr = df[df["Estado"] == "Aprobado"]
            for idx, row in df_apr.iterrows(): render_tarjeta(row, "apr")

    # --- PESTAÑA: RECHAZADOS ---
    with tab_rechazados:
        if not df.empty:
            df_rec = df[df["Estado"] == "Rechazado"]
            for idx, row in df_rec.iterrows(): render_tarjeta(row, "rec")

    # --- PESTAÑA: HISTORIAL Y DESCARGA EXCEL ---
    with tab_todos:
        st.write("### Base de datos completa")
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        col_m1.metric("⏳ Pendientes", total_pen)
        col_m2.metric("✅ Aprobados", total_apr)
        col_m3.metric("❌ Rechazados", total_rec)
        col_m4.metric("📊 Total Registros", total_reg)
        
        st.markdown("<br>", unsafe_allow_html=True)
        if not df.empty:
            excel_bytes = generar_excel_bytes(df)
            st.download_button(
                label=f"📥 Descargar Base de Datos Excel ({len(df)} filas)", data=excel_bytes,
                file_name="Base_Leche_LIF.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary"
            )
            st.markdown("<br>", unsafe_allow_html=True)
            cf1, cf2 = st.columns(2)
            with cf1: prov_filtro = st.selectbox("Filtrar por Proveedor:", ["Todos"] + list(df["Proveedor"].unique()), key="filter_prov")
            with cf2: fecha_filtro = st.selectbox("Filtrar por Fecha:", ["Todas"] + list(df["Fecha"].unique()), key="filter_fecha")
            
            df_mostrar = df
            if prov_filtro != "Todos": df_mostrar = df_mostrar[df_mostrar["Proveedor"] == prov_filtro]
            if fecha_filtro != "Todas": df_mostrar = df_mostrar[df_mostrar["Fecha"] == fecha_filtro]
            
            if df_mostrar.empty: st.warning("No hay registros con esos filtros.")
            else:
                for idx, row in df_mostrar.iterrows(): render_tarjeta(row, "tod")
